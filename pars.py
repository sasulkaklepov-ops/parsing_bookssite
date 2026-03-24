import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url="https://books.toscrape.com/"

response=requests.get(url)
html=response.text

soup=BeautifulSoup(html, "html.parser")


wb=Workbook()
ws=wb.active
ws.title="Книги"
ws["A1"]="Название книги"
ws["B1"]="Цена книги"
books=soup.find_all("article", class_="product_pod")
a=2
for book in books:
    t=book.find('a', title=True)
    title=t.get('title')
    pr=book.find("p", class_="price_color").text
    #print(title, ": ", pr)
    ws[f"A{a}"]=title
    ws[f"B{a}"]=pr
    a+=1
    wb.save("books.xlsx")
